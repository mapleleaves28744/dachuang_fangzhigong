import html
import io
import json
import os
import re
import zipfile
from typing import Dict, List


_LFS_POINTER_PREFIX = b"version https://git-lfs.github.com/spec/v1"
_XML_NAMESPACE_PATTERN = re.compile(r"\{[^}]+\}")
_HTML_TAG_PATTERN = re.compile(r"<[^>]+>")
_WHITESPACE_PATTERN = re.compile(r"\s+")
_PDF_TEXT_PATTERN = re.compile(rb"\(([^()]{1,240})\)\s*Tj")
_PDF_ARRAY_TEXT_PATTERN = re.compile(rb"\[(.*?)\]\s*TJ", re.DOTALL)
_PDF_ARRAY_SEGMENT_PATTERN = re.compile(rb"\(([^()]*)\)")
_PRINTABLE_STRING_PATTERN = re.compile(rb"[\x20-\x7e]{4,}")
_METADATA_ONLY_PREFIXES = ("文件名:", "MIME:", "大小:", "类型:", "来源:")
_TEXT_FILE_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".json",
    ".csv",
    ".tsv",
    ".log",
    ".py",
    ".js",
    ".ts",
    ".jsx",
    ".tsx",
    ".html",
    ".htm",
    ".xml",
    ".yaml",
    ".yml",
}


def is_git_lfs_pointer_bytes(raw_bytes: bytes) -> bool:
    return bool(raw_bytes and raw_bytes.startswith(_LFS_POINTER_PREFIX))


def is_git_lfs_pointer_file(path: str) -> bool:
    target = str(path or "").strip()
    if not target or not os.path.exists(target) or not os.path.isfile(target):
        return False
    try:
        with open(target, "rb") as file_obj:
            return is_git_lfs_pointer_bytes(file_obj.read(256))
    except Exception:
        return False


def normalize_text_preview(text: str, limit: int = 220) -> str:
    compact = _WHITESPACE_PATTERN.sub(" ", str(text or "")).strip()
    if len(compact) <= int(limit or 220):
        return compact
    return compact[: max(1, int(limit or 220))] + "..."


def _normalize_whitespace(text: str) -> str:
    lines = []
    for raw_line in str(text or "").replace("\r", "\n").split("\n"):
        cleaned = _WHITESPACE_PATTERN.sub(" ", raw_line).strip()
        if cleaned:
            lines.append(cleaned)
    return "\n".join(lines).strip()


def _decode_text_bytes(raw_bytes: bytes) -> str:
    if not raw_bytes:
        return ""

    for encoding in ("utf-8", "utf-8-sig", "gb18030", "gbk", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            normalized = _normalize_whitespace(text)
            if normalized:
                return normalized
        except Exception:
            continue
    return ""


def _strip_html_tags(raw_html: str) -> str:
    text = re.sub(r"<(script|style)\b[\s\S]*?</\1>", " ", str(raw_html or ""), flags=re.IGNORECASE)
    text = _HTML_TAG_PATTERN.sub(" ", text)
    text = html.unescape(text)
    return _normalize_whitespace(text)


def _extract_docx_text(raw_bytes: bytes) -> str:
    if not raw_bytes:
        return ""
    try:
        from xml.etree import ElementTree

        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
            document_xml = archive.read("word/document.xml")
        root = ElementTree.fromstring(document_xml)
        fragments = []
        for element in root.iter():
            tag = _XML_NAMESPACE_PATTERN.sub("", str(element.tag or ""))
            if tag == "t":
                text = str(element.text or "").strip()
                if text:
                    fragments.append(text)
        return _normalize_whitespace(" ".join(fragments))
    except Exception:
        return ""


def _extract_xlsx_text(raw_bytes: bytes) -> str:
    if not raw_bytes:
        return ""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        fragments = []
        for sheet in workbook.worksheets:
            if sheet.title:
                fragments.append(f"[工作表]{sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if str(cell or "").strip()]
                if cells:
                    fragments.append(" | ".join(cells))
        return _normalize_whitespace("\n".join(fragments))
    except Exception:
        return ""


def _decode_pdf_fragment(raw_bytes: bytes) -> str:
    text = raw_bytes.decode("latin-1", errors="ignore")
    text = text.replace(r"\(", "(").replace(r"\)", ")").replace(r"\n", "\n").replace(r"\r", "")
    return _normalize_whitespace(text)


def _extract_pdf_text(raw_bytes: bytes) -> str:
    if not raw_bytes:
        return ""

    fragments = []
    for match in _PDF_TEXT_PATTERN.findall(raw_bytes):
        decoded = _decode_pdf_fragment(match)
        if decoded:
            fragments.append(decoded)

    if not fragments:
        for block in _PDF_ARRAY_TEXT_PATTERN.findall(raw_bytes):
            parts = []
            for match in _PDF_ARRAY_SEGMENT_PATTERN.findall(block):
                decoded = _decode_pdf_fragment(match)
                if decoded:
                    parts.append(decoded)
            if parts:
                fragments.append(" ".join(parts))

    if not fragments:
        rough_strings = []
        for item in _PRINTABLE_STRING_PATTERN.findall(raw_bytes):
            decoded = _decode_pdf_fragment(item)
            if decoded and len(decoded) >= 8:
                rough_strings.append(decoded)
        fragments = rough_strings

    deduped = []
    seen = set()
    for fragment in fragments:
        key = fragment.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(fragment)
        if len("\n".join(deduped)) >= 18000:
            break

    return _normalize_whitespace("\n".join(deduped))


def _looks_like_html(name: str, mime: str, text: str) -> bool:
    lower_name = str(name or "").lower()
    mime_text = str(mime or "").lower()
    raw = str(text or "")
    return (
        lower_name.endswith((".html", ".htm"))
        or "html" in mime_text
        or ("<html" in raw.lower() or "<body" in raw.lower())
    )


def _looks_like_json(name: str, mime: str, text: str) -> bool:
    lower_name = str(name or "").lower()
    mime_text = str(mime or "").lower()
    raw = str(text or "").strip()
    return (
        lower_name.endswith(".json")
        or "json" in mime_text
        or (raw.startswith("{") and raw.endswith("}"))
        or (raw.startswith("[") and raw.endswith("]"))
    )


def _format_json_for_text(raw: str) -> str:
    try:
        parsed = json.loads(raw)
        return _normalize_whitespace(json.dumps(parsed, ensure_ascii=False, indent=2))
    except Exception:
        return _normalize_whitespace(raw)


def _is_metadata_only_summary(text: str) -> bool:
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    if not lines:
        return False
    return all(line.startswith(_METADATA_ONLY_PREFIXES) for line in lines[:4])


def extract_text_from_learning_asset(
    *,
    name: str = "",
    mime: str = "",
    content: str = "",
    summary: str = "",
    file_bytes: bytes = b"",
) -> Dict[str, object]:
    provided_content = _normalize_whitespace(content)
    summary_text = _normalize_whitespace(summary)
    lower_name = str(name or "").lower()
    mime_text = str(mime or "").lower()

    extracted_text = ""
    method = "empty"
    warnings: List[str] = []

    if file_bytes and is_git_lfs_pointer_bytes(file_bytes):
        warnings.append("uploaded_file_is_git_lfs_pointer")

    if file_bytes:
        if lower_name.endswith(".docx") or "wordprocessingml.document" in mime_text:
            extracted_text = _extract_docx_text(file_bytes)
            method = "docx_xml" if extracted_text else method
        elif lower_name.endswith(".xlsx") or "spreadsheetml.sheet" in mime_text:
            extracted_text = _extract_xlsx_text(file_bytes)
            method = "xlsx_cells" if extracted_text else method
        elif lower_name.endswith(".pdf") or "pdf" in mime_text:
            extracted_text = _extract_pdf_text(file_bytes)
            method = "pdf_text" if extracted_text else method
        elif lower_name.endswith(tuple(_TEXT_FILE_EXTENSIONS)) or mime_text.startswith("text/") or mime_text in {
            "application/json",
            "application/xml",
            "text/csv",
            "text/tab-separated-values",
        }:
            decoded = _decode_text_bytes(file_bytes)
            if decoded:
                if _looks_like_html(lower_name, mime_text, decoded):
                    extracted_text = _strip_html_tags(decoded)
                    method = "html_text"
                elif _looks_like_json(lower_name, mime_text, decoded):
                    extracted_text = _format_json_for_text(decoded)
                    method = "json_text"
                else:
                    extracted_text = decoded
                    method = "text_decode"
        else:
            decoded = _decode_text_bytes(file_bytes[:12000])
            if decoded and len(decoded) >= 20:
                extracted_text = decoded
                method = "binary_best_effort"

    candidates = []
    if provided_content:
        candidates.append(("provided_content", provided_content))
    if extracted_text:
        candidates.append((method, extracted_text))
    if summary_text and not _is_metadata_only_summary(summary_text):
        candidates.append(("summary", summary_text))

    primary_method = "empty"
    primary_text = ""
    if candidates:
        primary_method, primary_text = max(candidates, key=lambda item: len(str(item[1] or "")))

    if primary_method == "summary" and extracted_text:
        primary_method = method or "file_extract"
        primary_text = extracted_text

    if not primary_text and summary_text:
        primary_text = summary_text
        primary_method = "summary_only"

    chunks = split_text_into_chunks(primary_text, chunk_size=720, overlap=120, max_chunks=24) if primary_text else []
    quality = "empty"
    if primary_text:
        if len(primary_text) >= 120 and len(chunks) >= 1:
            quality = "rich_text"
        elif len(primary_text) >= 24:
            quality = "light_text"
        else:
            quality = "minimal_text"

    return {
        "text": primary_text,
        "preview": normalize_text_preview(primary_text, 220),
        "method": primary_method,
        "warnings": warnings,
        "content_length": len(primary_text),
        "chunk_count": len(chunks),
        "quality": quality,
    }


def split_text_into_chunks(
    text: str,
    *,
    chunk_size: int = 700,
    overlap: int = 120,
    max_chunks: int = 48,
) -> List[str]:
    normalized = _normalize_whitespace(text)
    if not normalized:
        return []

    size = max(120, int(chunk_size or 700))
    overlap_size = max(0, min(size // 2, int(overlap or 120)))
    limit = max(1, int(max_chunks or 48))

    if len(normalized) <= size:
        return [normalized]

    segments = []
    buffer = ""
    pieces = [piece.strip() for piece in re.split(r"(?<=[。！？；.!?;])|\n+", normalized) if piece.strip()]
    for piece in pieces:
        candidate = f"{buffer}{piece}" if buffer else piece
        if len(candidate) <= size:
            buffer = candidate + ("\n" if len(candidate) < size else "")
            continue
        if buffer:
            segments.append(buffer.strip())
            if len(segments) >= limit:
                return segments
            tail = buffer[-overlap_size:] if overlap_size else ""
            buffer = f"{tail}{piece}".strip()
        else:
            start = 0
            while start < len(piece):
                end = min(len(piece), start + size)
                chunk = piece[start:end].strip()
                if chunk:
                    segments.append(chunk)
                    if len(segments) >= limit:
                        return segments
                if end >= len(piece):
                    buffer = ""
                    break
                start = max(end - overlap_size, start + 1)

    if buffer.strip() and len(segments) < limit:
        segments.append(buffer.strip())

    return segments[:limit]
